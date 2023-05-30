import datetime
import os.path
from flask import current_app, send_file, make_response
from flask.views import MethodView
from flask_smorest import Blueprint
from flask_jwt_extended import jwt_required
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from ..db import db
from ..models import ItemModel
from ..models.transactions.inventory_balances import InventoryBalancesModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

blp = Blueprint("Reports", __name__, description="Report Handling")


@blp.route("/stockholding")
class StockHoldingReport(MethodView):
    blp.response(200, "Download success")

    @jwt_required(fresh=True)
    def get(self):
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        wb = Workbook()
        inventory_summary_query = db.session.query(
            InventoryBalancesModel.item_id,
            func.avg(InventoryBalancesModel.unit_cost).label('average_unit_cost'),
            func.sum(InventoryBalancesModel.quantity).label('total_quantity')
        ).group_by(InventoryBalancesModel.item_id).all()

        ws = wb.active
        ws.title = "stock_holding_report"
        heading = ["itemID", "Item_name", "unit_cost", "total_quantity", "total_cost"]
        light_blue_fill = PatternFill(start_color="C6E2FF", end_color="C6E2FF", fill_type="solid")
        thin_border = Border(bottom=Side(style='thick'), right=Side(style='thin'))
        ws.append(heading)
        lines = []
        for balance in inventory_summary_query:
            item = ItemModel.query.get(balance[0])
            cost = balance[1]*balance[2]
            item_line = [balance[0], item.item_name, balance[1], balance[2], cost]
            ws.append(item_line)
            lines.append(item_line)

        for col in range(1, len(heading) + 1):
            ws[get_column_letter(col) + '1'].font = Font(bold=True, color="000000", italic=True)
            ws[get_column_letter(col) + '1'].fill = light_blue_fill
            ws[get_column_letter(col) + '1'].border = thin_border
        file_path = os.path.join(current_app.static_folder, f'reports/custom/stockholding_{today}.xlsx')
        wb.save(file_path)
        filename = f"Stock-holding-{today}.xlsx"
        response = make_response(send_file(file_path))
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return response
